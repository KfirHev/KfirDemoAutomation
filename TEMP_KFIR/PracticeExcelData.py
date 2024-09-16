import openpyxl
from openpyxl import Workbook, worksheet

path_to_excell = r'/TestData/ExcelDataToPytest.xlsx'

book = openpyxl.load_workbook(path_to_excell)
sheet = book.active
cell = sheet.cell(row=1, column=2)
# sheet.cell(row=2,column=2).value = 'KFIRS'
book.save(path_to_excell)

test_list = []

for rw in range(2, sheet.max_row + 1):
    test_data = {}
    for cl in range(2, sheet.max_column + 1):
        test_data[sheet.cell(row=1, column=cl).value] = sheet.cell(row=rw, column=cl).value
    test_list.append(test_data)



print(test_list)

print("FOR BRANCH TESTING ")
print ("Add from another user for MERGE TEST")


# print(sheet.max_column)
# print(sheet.max_row)
#
# print(sheet.cell(row=2,column=2).value)
# print(sheet['D2'].value)
# print(cell.value)
# print(book)

# printing values of 'TestCase2 in the excell
# for rw in range(1, sheet.max_row + 1):
#     if sheet.cell(row=rw, column=1).value == 'TestCase2':
#         for cl in range(2, sheet.max_column + 1):
#             Test2data.update({sheet.cell(row=1, column=cl).value: sheet.cell(row=rw, column=cl).value})
