import openpyxl


class HomePageData:
    test_home_page_login = [
        {'user_name_hint': 'Username',
         'pw_hint': 'Password',
         'users': {'standard_user', 'problem_user', 'performance_glitch_user', 'error_user',
                   'visual_user'},
         'invalid_users': ('Darth_Vaders_Nephew', ' standard_users', ' standard_user', 'standard_user '),
         'locked_user': 'locked_out_user',
         'password': 'secret_sauce',
         'invalid password': ('bs password', '*secret_sauce', ' secret_sauce', 'secret_sauce '),
         'empty': ''
         }]

    # @staticmethod
    # def get_data(self, request):
    #     return request.param

    # @staticmethod
    # def get_data_excel(test_case):
    #
    #     #path_to_excell = r'C:\Users\hkfir\PycharmProjects\pythonKfirFramework\TestData\ExcelDataToPytest.xlsx'
    #     path_to_excell = '/app/TestData/ExcelDataToPytest.xlsx'  # TODO remove app/ when running locally
    #     book = openpyxl.load_workbook(path_to_excell)
    #     sheet = book.active
    #
    #     test_list = []
    #
    #     for rw in range(1, sheet.max_row + 1):
    #         test_data = {}
    #         if sheet.cell(row=rw, column=1).value == test_case:
    #             for cl in range(2, sheet.max_column + 1):
    #                 test_data[sheet.cell(row=1, column=cl).value] = sheet.cell(row=rw, column=cl).value
    #             test_list.append(test_data)
    #     return test_list
